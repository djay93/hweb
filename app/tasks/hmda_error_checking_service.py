from datetime import datetime
import pandas as pd
from typing import Any, List, Dict, Set
from openpyxl import load_workbook
from app.core.error_handler import ErrorHandler
from app.models.enum import TaskStatus
from app.models import JobTask, ActivityLog
from app.extensions import db
import logging
from pathlib import Path
import csv
import json

logger = logging.getLogger(__name__)

class ErrorCheckingService:
    def __init__(self, job_task_id: str):
        self.job_task_id = job_task_id
        self.error_handler = ErrorHandler()
        self.replacements: Dict[str, str] = {}
        self.valid_separators: Set[str] = {',', ';', '|'}


    def process_error_checking(
        self,
        input_file_path: str,
        sheet_name: str = 'last-sheet',
        target_columns: List[str] = None,
        replacer_file_path: str = None,
        remove_empty_rows: bool = False,
        sort_columns: List[str] = None
    ):
        
        try:
            # Load replacements if file provided
            if replacer_file_path:
                self.replacements = self._load_replacements(replacer_file_path)
            
            # Validate file path
            input_file_path = Path(input_file_path)
            if not input_file_path.exists():
                raise FileNotFoundError(f"File not found: {input_file_path}")
            
            # Read the Excel file
            logger.info(f"Processing file: {input_file_path}")
            df_dict = pd.read_excel(input_file_path, sheet_name=None)

            # Get target sheet
            sheet_key = self._get_sheet_key(df_dict, sheet_name)
            df = df_dict[sheet_key]

            # Get and validate target columns
            target_columns = self._get_target_columns(df, target_columns)
            if not target_columns:
                raise ValueError("No valid target columns found in sheet")
            
            # Process each target column
            logger.info(f"Processing {len(target_columns)} columns in sheet '{sheet_key}'")
            for column in target_columns:
                logger.debug(f"Processing column: {column}")
                df[column] = df[column].apply(self._process_cell)
            logger.info(f"Successfully processed file with {len(df)} rows")

            # Remove empty rows if requested
            if remove_empty_rows:
                logger.info(f"Removing empty rows from target columns: {target_columns}")
                original_count = len(df)
                df = df.dropna(subset=target_columns, how='all')
                removed_count = original_count - len(df)
                logger.info(f"Removed {removed_count} empty rows. Remaining rows: {len(df)}")


            # Sort by specified columns if provided
            if sort_columns:
                valid_sort_columns = [col for col in sort_columns if col in df.columns]
                if valid_sort_columns:
                    logger.info(f"Sorting by columns: {valid_sort_columns}")
                    df = df.sort_values(by=valid_sort_columns, na_position='last')
                else:
                    logger.warning(f"No valid sort columns found in: {sort_columns}")

            # Export to csv
            csv_file_path = self._export_to_csv(df, input_file_path)

            # Export to excel
            excel_file_path = self._create_processed_excel(
                original_file=input_file_path,
                df=df,
                sheet_key=sheet_key,
                target_columns=target_columns
            )

            # Update job task status
            self._update_job_task_status(TaskStatus.SUCCESS.name)

            return csv_file_path, excel_file_path

        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            error = self.error_handler.handle_error(
                error=e,
                source_id=self.job_task_id,
                source_type="job_tasks",
                context={
                    "service": self.__class__.__name__,
                    "method": self.process_error_checking.__name__,
                    "file_path": str(input_file_path),
                    "sheet_name": sheet_name
                }
            )
            self._update_job_task_status(TaskStatus.FAILED.name, error.to_dict())
            raise

    def _process_cell(self, value: Any) -> str:
        """Process individual cell value with error handling and validation."""
        try:
            # Handle null values
            if pd.isna(value):
                return value
            
            value = str(value).strip()
            if not value:
                return value
            
            # Determine separator
            separator = next((sep for sep in self.valid_separators if sep in value), None)
            
            if separator:
                parts = [part.strip() for part in value.split(separator)]
                processed_parts = [self._apply_replacements(part) for part in parts]
                return separator.join(processed_parts)
            
            return self._apply_replacements(value)
            
        except Exception as e:
            logger.warning(f"Error processing cell value '{value}': {str(e)}")
            return value  # Return original value on error
    
    def _apply_replacements(self, text: str) -> str:
        """Apply text replacements with error handling."""
        try:
            text = text.strip()
            for old_text, new_text in self.replacements.items():
                text = text.replace(old_text, new_text)
            return text
        except Exception as e:
            logger.warning(f"Error applying replacements to '{text}': {str(e)}")
            return text

    def _load_replacements(self, file_path: str) -> Dict[str, str]:
        """Load text replacements from configuration file."""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Replacement file not found: {file_path}")

            if file_path.suffix.lower() == '.json':
                with open(file_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            elif file_path.suffix.lower() == '.csv':
                replacements = {}
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        replacements[row['old_text']] = row['new_text']
                return replacements
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
        except Exception as e:
            logger.error(f"Error loading replacements from {file_path}: {str(e)}")
            raise

    @staticmethod
    def _get_target_columns(df: pd.DataFrame, target_columns: List[str]) -> List[str]:
        """Get list of columns to process."""
        available_columns = set(df.columns)
        valid_columns = [col for col in target_columns if col in available_columns]
        return valid_columns

    @staticmethod
    def _get_sheet_key(df_dict: Dict, sheet_name: str) -> str:
        """Determine which sheet to use."""
        if sheet_name == 'last-sheet':
            return list(df_dict.keys())[-1]
        elif sheet_name == 'first-sheet':
            return list(df_dict.keys())[0]
        elif sheet_name in df_dict:
            return sheet_name
        raise ValueError(f"Sheet '{sheet_name}' not found")

    def _validate_data(self, data: pd.DataFrame) -> pd.DataFrame:
        pass

    def _save_results(self, results: pd.DataFrame, job_task_id: str):
        pass

    def _update_job_task_status(self, status: str, error_message: str = None):
        task = JobTask(id=self.job_task_id, status=status, completed_at=datetime.now(), retries=0)
        activity_log = ActivityLog(job_task_id=self.job_task_id, activity_type=status, status=status, error_message=error_message)
            
        db.session.merge(task)
        db.session.add(activity_log)
        db.session.commit()

    def _log_error(self, message: str):
        pass

    def _create_processed_excel(
        self,
        original_file: Path,
        df: pd.DataFrame,
        sheet_key: str,
        target_columns: List[str]
    ) -> Path:
        try:
            # Create output path
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_excel = original_file.parent / f"{original_file.stem}_processed_{timestamp}{original_file.suffix}"
            
            # Load original workbook
            workbook = load_workbook(original_file)
            worksheet = workbook[sheet_key]
            
            # Get all column headers and their positions
            headers = {}
            for cell in worksheet[1]:
                if cell.value:  # Skip empty headers
                    headers[cell.value] = cell.column_letter
            
            # Write all data from the cleaned DataFrame
            for row_idx, df_row in enumerate(df.itertuples(), start=2):
                for col_name in df.columns:
                    if col_name in headers:
                        cell = worksheet[f"{headers[col_name]}{row_idx}"]
                        original_format = cell._style
                        cell.value = getattr(df_row, col_name)
                        cell._style = original_format
            
            # Clear any remaining rows in the worksheet
            max_row = worksheet.max_row
            if max_row > len(df) + 1:  # +1 for header row
                for row in range(len(df) + 2, max_row + 1):
                    for col in headers.values():
                        cell = worksheet[f"{col}{row}"]
                        cell.value = None
            
            # Save processed workbook
            workbook.save(output_excel)
            logger.info(f"Created processed Excel file: {output_excel}")
            
            return output_excel
            
        except Exception as e:
            logger.error(f"Failed to create processed Excel file: {str(e)}")
            error = self.error_handler.handle_error(
                error=e,
                source_id=self.job_task_id,
                source_type="job_tasks",
                context={
                    "service": self.__class__.__name__,
                    "method": self._create_processed_excel.__name__,
                    "file_path": str(original_file)
                }
            )
            raise

    def _export_to_csv(self, df: pd.DataFrame, file_path: Path) -> Path:
        """
        Export DataFrame to CSV with timestamp in filename.
        
        Args:
            df: DataFrame to export
            file_path: Original file path to determine output directory
            
        Returns:
            Path: Path to the created CSV file
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = Path(file_path).parent / f"{file_path.stem}_{timestamp}.csv"
            
            df.to_csv(
                output_path,
                index=False,
                encoding='utf-8-sig',
                quoting=csv.QUOTE_ALL
            )
            logger.info(f"Exported processed data to: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to export CSV: {str(e)}")
            raise
